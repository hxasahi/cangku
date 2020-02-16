# -*- coding: cp936 -*-
'''   制作：韩旭
      时间：2020.2.15 19:30
      版本：仓库管理系统v1.0   '''
      #项目状态：未完

from openpyxl import load_workbook
wb = load_workbook('F:\pyadmin\products.xlsx')
ws = wb.active
#查
print('')
print('请输入ID查询：')
shangpin = input()
nums = ws.max_row
for i in range(2,nums+1):
    if shangpin == ws.cell(row=i,column=1).value:
        print('名称：'),
        print(ws.cell(row=i,column=2).value)
        print('价格：{}'.format(ws.cell(row=i,column=3).value))
#增
tianjia=[]
print('添加ID:')
zancun=int(input())
tianjia.append(zancun)
print('添加名称')
zancun=raw_input()
tianjia.append(zancun)
print('添加价格')
zancun=int(input())
tianjia.append(zancun)
ws.append(tianjia)
wb.save('F:\pyadmin\products.xlsx')
#删
print('')
shanchuid=int(input('请输入删除商品ID:'))
shan=['','','']
nums = ws.max_row
for i in range(2,nums+1):
    if shanchuid == ws.cell(row=i,column=1).value and i != nums :
        ws.cell(row=i,column=1).value=ws.cell(row=nums,column=1).value
        ws.cell(row=i,column=2).value=ws.cell(row=nums,column=2).value
        ws.cell(row=i,column=3).value=ws.cell(row=nums,column=3).value
        ws.cell(row=nums,column=1).value=''
        ws.cell(row=nums,column=2).value=''
        ws.cell(row=nums,column=3).value=''
        wb.save('F:\pyadmin\products.xlsx')
    elif shanchuid == ws.cell(row=i,column=1).value and i == nums :
        ws.cell(row=i,column=1).value=''
        ws.cell(row=i,column=2).value=''
        ws.cell(row=i,column=3).value=''
        wb.save('F:\pyadmin\products.xlsx')
#改
print('')
xiugaiid=int(input('请输入修改商品ID'))
nums = ws.max_row
for i in range(2,nums+1):
    if xiugaiid == ws.cell(row=i,column=1).value and i != nums :
        xiuid=int(input('请输入修改后商品ID：'))
        ws.cell(row=i,column=1).value=xiuid
        xiuname=raw_input('请输入修改后商品名称：')
        ws.cell(row=i,column=2).value=xiuname
        xiujiage=int(input('请输入修改后商品价格：'))
        ws.cell(row=i,column=3).value=xiujiage
        wb.save('F:\pyadmin\products.xlsx')
