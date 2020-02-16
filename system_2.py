# -*- coding: cp936 -*-
'''   制作：韩旭
      时间：2020.2.16 19:54
      版本：仓库管理系统v1.1   '''
      #项目状态：未完

from openpyxl import load_workbook
wb = load_workbook('F:\pyadmin\products.xlsx')
ws = wb.active
#查
def chaxun():
    print('')
    print('请输入ID查询：')
    productid = input()
    nums = ws.max_row
    for i in range(2,nums+1):
        if productid == ws.cell(row=i,column=1).value:
            print('名称：'),
            print(ws.cell(row=i,column=2).value)
            print('')
            print('价格：{}'.format(ws.cell(row=i,column=3).value))
#增
def addproduct():
    tianjia=[]
    print('添加ID:')
    zancun=int(input())
    tianjia.append(zancun)
    print('添加名称')
    zancun=raw_input()
    tianjia.append(zancun)
    print('添加价格')
    zancun=int(input())
    tianjia.append(zancun)
    ws.append(tianjia)
    wb.save('F:\pyadmin\products.xlsx')
#删
def delete():
    print('')
    shanchuid=int(input('请输入删除商品ID:'))
    shan=['','','']
    nums = ws.max_row
    for i in range(2,nums+1):
        if shanchuid == ws.cell(row=i,column=1).value and i != nums :
            ws.cell(row=i,column=1).value=ws.cell(row=nums,column=1).value
            ws.cell(row=i,column=2).value=ws.cell(row=nums,column=2).value
            ws.cell(row=i,column=3).value=ws.cell(row=nums,column=3).value
            ws.cell(row=nums,column=1).value=''
            ws.cell(row=nums,column=2).value=''
            ws.cell(row=nums,column=3).value=''
            wb.save('F:\pyadmin\products.xlsx')
        elif shanchuid == ws.cell(row=i,column=1).value and i == nums :
            ws.cell(row=i,column=1).value=''
            ws.cell(row=i,column=2).value=''
            ws.cell(row=i,column=3).value=''
            wb.save('F:\pyadmin\products.xlsx')
#改
def correct():
    print('')
    xiugaiid=int(input('请输入修改商品ID'))
    nums = ws.max_row
    for i in range(2,nums+1):
        if xiugaiid == ws.cell(row=i,column=1).value and i != nums :
            xiuid=int(input('请输入修改后商品ID：'))
            ws.cell(row=i,column=1).value=xiuid
            xiuname=raw_input('请输入修改后商品名称：')
            ws.cell(row=i,column=2).value=xiuname
            xiujiage=int(input('请输入修改后商品价格：'))
            ws.cell(row=i,column=3).value=xiujiage
            wb.save('F:\pyadmin\products.xlsx')
print('')
print('')
print('                          欢迎使用仓库管理系统')
print('')
print('1 查询        '),
print('2 增加      '),
print('3 删除   '),
print('4 修改     ')
print('')
print('99 退出')
n = input('               请按下对应数字选择：')
while 1 :
    if n == 1 :
        chaxun()
    elif n == 2 :
        addproduct()
    elif n == 3 :
        delete()
    elif n == 4 :
        correct()
    elif n == 99 :
        break
    #n = input('               请按下对应数字选择：')
        
    else :
        print('')
        print('               请输入合法字符!')
        
    print('')
    print('               1 继续    '),
    print('99 退出')
    print('')

    c = input('               是否继续？')
    
    while c != 1 and c != 99 :
        print('')
        print('               请输入合法字符!')
        print('')
        print('               1 继续    '),
        print('99 退出')
        print('')
        c = input('               是否继续？')
    if c == 1 :
        print('')
        print('1 查询        '),
        print('2 增加      '),
        print('3 删除   '),
        print('4 修改     ')
        print('')
        print('99 退出')
        print('')     
    elif c == 99 :
        break
    
    n = input('               请按下对应数字选择：')       
